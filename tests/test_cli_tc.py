import json

import pytest

from qatc.cli import main
from qatc.knowledge.models import SlotStatus
from qatc.knowledge.store import KnowledgeStore


@pytest.fixture()
def ready(cfg_env):
    main(["slot", "init", "파티편성", "--game", "starrail", "--code", "PARTY"])
    main(["slot", "set", "파티편성", "core_action", "--status", "filled",
          "--value", "파티를 짜고 적용한다"])
    return cfg_env


def _payload(sub="정상 동작", **over):
    """유효한 `tc add` 페이로드 하나.

    첫 인자 이름이 예전엔 `title` 이었다. 이 파일의 많은 테스트가 저장된 TC를
    구별하는 데 그 값을 썼는데, `tc add` 가 더 이상 `title` 을 읽지 않으므로
    (표 재설계로 소분류가 그 자리를 대신한다) 같은 역할을 `sub`(소분류)가
    잇는다 — 호출부(`_payload("파일에서 읽은 TC")` 처럼 위치 인자로 부르는
    자리 포함)를 그대로 두기 위해 이름만 바꿨다.
    """
    item = {"middle": "파티 편성", "sub": sub,
            "precondition": "파티 편성 화면", "steps": ["파티 적용을 누른다"],
            "expected": ["파티가 적용된다"], "rationale": "core_action 슬롯에서 도출"}
    item.update(over)
    return json.dumps({"testcases": [item]}, ensure_ascii=False)


def test_plan_lists_filled_family(ready, capsys):
    assert main(["tc", "plan", "파티편성", "--json"]) == 0
    data = json.loads(capsys.readouterr().out)
    assert "정상 경로" in [p["family"] for p in data["planned"]]


def test_plan_lists_skipped_with_reason(ready, capsys):
    main(["tc", "plan", "파티편성", "--json"])
    data = json.loads(capsys.readouterr().out)
    skipped = {s["family"]: s for s in data["skipped"]}
    assert skipped["재화 부족"]["status"] == "empty"
    assert skipped["재화 부족"]["slot"] == "cost"


def test_add_accepts_planned_family(stdin_text, ready, capsys, monkeypatch):
    stdin_text(_payload())
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 0
    with KnowledgeStore(ready / "starrail.db") as s:
        # `title` 은 더 이상 읽지 않는다 — 저장된 TC를 식별하는 값은
        # `category_sub`(소분류)로 옮겨왔다.
        assert [t.category_sub for t in s.testcases("파티편성")] == ["정상 동작"]


def test_add_rejects_unplanned_family(stdin_text, ready, capsys, monkeypatch):
    stdin_text(_payload())
    rc = main(["tc", "add", "파티편성", "--family", "재화 부족",
               "--origin", "inferred", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "재화 부족" in out
    assert "cost" in out
    assert "tc plan" in out


def test_add_rejects_unknown_family(stdin_text, ready, monkeypatch, capsys):
    """등록되지 않은 계열 이름은 `tc add` 에서 거부된다."""
    stdin_text(_payload())
    rc = main(["tc", "add", "파티편성", "--family", "없는계열",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    assert "등록되지 않은 계열" in capsys.readouterr().out


def _inject_unregistered_family_slot(root):
    """CLI 검증을 우회해 미등록 계열 슬롯을 DB에 직접 넣는다.

    `slot add --family` 의 검증은 CLI 경계에만 있으므로, 저장소 API 를 직접
    부르거나 그 검증이 생기기 전에 만들어진 DB 를 열면 이 상태가 실제로 존재한다.
    """
    with KnowledgeStore(root / "starrail.db") as s:
        s.add_slot("파티편성", "네트워크", "통신이 끊기면", "중단됨")
        s.set_slot("파티편성", "네트워크", SlotStatus.FILLED, "전투 중 통신이 끊긴다")


def test_plan_and_add_agree_on_unregistered_family(stdin_text, ready, monkeypatch, capsys):
    """`tc plan` 과 `tc add` 가 미등록 계열에 대해 같은 답을 해야 한다.

    예전에는 `tc plan` 이 `중단됨`(오타)을 **`정상` / Medium 으로 계획**하고
    `tc add` 가 그대로 받아, 최종 xlsx 의 `정상 경로` 칸에 근거 없는 TC가
    들어갔다. 오타는 그 칸에서 보이지 않는다.
    """
    _inject_unregistered_family_slot(ready)
    capsys.readouterr()          # 앞선 명령의 확인 문구를 버린다

    assert main(["tc", "plan", "파티편성", "--json"]) == 0
    data = json.loads(capsys.readouterr().out)
    assert "중단됨" not in [p["family"] for p in data["planned"]]
    skipped = {s["family"]: s for s in data["skipped"]}
    assert "등록되지 않은 계열" in skipped["중단됨"]["reason"]

    stdin_text(_payload())
    rc = main(["tc", "add", "파티편성", "--family", "중단됨",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "등록되지 않은 계열" in out
    assert _stored(ready) == []


def test_add_rejects_missing_required_field(stdin_text, ready, monkeypatch, capsys):
    """필수 필드가 두 개 빠지면 실제 거부 메시지에 **둘 다** 나와야 한다.

    예전 판은 `rc == 1` 과 `"steps" in out` 만 봤고, 그 두 단언은 **검증이 0인
    구현에서도 둘 다 참**이었다 — 검증 블록을 통째로 지우면 `item["steps"]` 에서
    `KeyError: 'steps'` 가 나고 `cli.py` 의 범용 핸들러가 그것을
    `오류: KeyError: 'steps'` 로 출력하며 rc=1 을 돌려주기 때문이다.
    (실측: 블록 삭제 후 `pytest tests/test_cli_tc.py` → 13 passed.)
    실제 메시지를 고정해 그 구멍을 막는다.
    """
    bad = json.dumps({"testcases": [{"middle": "파티 편성", "sub": "제목만 있음"}]},
                     ensure_ascii=False)
    stdin_text(bad)
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "필수 필드가 없습니다" in out
    assert "steps" in out
    assert "expected" in out
    assert "KeyError" not in out          # 날 예외가 아니라 우리 메시지여야 한다


def test_add_missing_field_message_names_only_what_is_missing(stdin_text, ready, monkeypatch, capsys):
    """빠진 필드만 나열한다.

    필수 필드 세 개를 항상 찍는 구현으로는 통과할 수 없어야, 메시지가 실제로
    무엇이 빠졌는지 계산한다는 것이 고정된다.
    """
    bad = json.dumps({"testcases": [{"middle": "m", "sub": "s", "steps": ["s"]}]},
                     ensure_ascii=False)
    stdin_text(bad)
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "expected" in out
    assert "steps" not in out             # 있는 필드는 언급하지 않는다


def _one(**over):
    """유효한 TC 항목 하나를 만들고 지정한 키만 덮어쓴다."""
    item = {
        "middle": "파티 편성",
        "sub": "정상 동작",
        "steps": ["파티 적용을 누른다"],
        "expected": ["파티가 적용된다"],
    }
    item.update(over)
    return json.dumps({"testcases": [item]}, ensure_ascii=False)


def _stored(root):
    with KnowledgeStore(root / "starrail.db") as s:
        return s.testcases("파티편성")


def test_add_rejects_string_steps(stdin_text, ready, monkeypatch, capsys):
    """`"steps": "한 줄"` 은 조용히 글자 단위로 쪼개지면 안 된다.

    이 명령의 호출자는 LLM 이고, 배열이어야 할 자리에 문자열을 주는 것은 가장
    흔한 JSON 형태 오류다. 예전에는 truthiness 검사만 통과하면 문자열을 순회해
    `['한', ' ', '줄']` 이 rc=0 + 성공 메시지와 함께 최종 xlsx 절차 칸까지 갔다.
    """
    stdin_text(_one(steps="한 줄"))
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "testcases[0].steps" in out      # 어느 필드가 틀렸는지 짚는다
    assert "배열" in out                     # 다음 조치
    assert _stored(ready) == []              # 쓰레기가 저장되지 않았다


def test_add_rejects_string_expected(stdin_text, ready, monkeypatch, capsys):
    stdin_text(_one(expected="한 줄"))
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    assert "testcases[0].expected" in capsys.readouterr().out
    assert _stored(ready) == []


def test_add_rejects_non_dict_item(stdin_text, ready, monkeypatch, capsys):
    """항목이 객체가 아니면 날 `AttributeError` 대신 다음 조치를 알린다."""
    bad = json.dumps({"testcases": [["제목만 든 배열"]]}, ensure_ascii=False)
    stdin_text(bad)
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "testcases[0]" in out
    assert "AttributeError" not in out      # 파이썬 타입명이 새어나오면 안 된다


def test_add_rejects_non_string_list_element(stdin_text, ready, monkeypatch, capsys):
    """배열 원소가 문자열이 아니면 `str()` 로 뭉개지 않고 거부한다."""
    stdin_text(_one(steps=[{"a": 1}]))
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    assert "testcases[0].steps[0]" in capsys.readouterr().out
    assert _stored(ready) == []


@pytest.mark.parametrize("raw, kind", [
    ('[{"title":"t","steps":["s"],"expected":["e"]}]', "list"),
    ('"testcases"', "str"),
    ('42', "int"),
    ('null', "NoneType"),
], ids=["array", "string", "number", "null"])
def test_add_rejects_non_object_top_level(stdin_text, ready, monkeypatch, capsys, raw, kind):
    """최상위가 객체가 아니면 날 `AttributeError` 대신 기대 형태를 알려준다.

    라운드 1a 는 항목 하나하나(`_validate_item`)만 검사하고 **페이로드 자체는
    아무도 검사하지 않았다.** 실측:

        $ echo '[{"title":"t"}]' > bad.json
        $ qatc tc add ... --json bad.json

        오류: AttributeError: 'list' object has no attribute 'get'

    (`오류:` 앞의 빈 줄까지 포함해서, `cli.py` 의 범용 예외 핸들러가 낸
    출력이다.) 이 명령의 호출자는 인터뷰를 진행하는 모델이라, 파이썬 타입명은
    다음에 무엇을 할지 알려주지 않는다.
    """
    stdin_text(raw)
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "AttributeError" not in out
    assert kind in out                        # 무엇이 왔는지 짚는다
    assert "testcases" in out                 # 기대 형태를 알려준다
    assert out.lstrip() == out                # 범용 핸들러가 내던 빈 줄이 없다
    assert _stored(ready) == []


def test_add_missing_testcases_key_names_next_action(stdin_text, ready, monkeypatch, capsys):
    """객체이긴 한데 `testcases` 가 없는 경우도 다음 조치를 함께 알린다."""
    stdin_text('{"cases": []}')
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "testcases" in out
    assert "middle" in out                    # 기대 형태를 예시로 보여준다


def test_add_rejects_unparseable_json_text(stdin_text, ready, monkeypatch, capsys):
    """JSON 자체가 깨진 경우 — 이미 막혀 있던 경로를 고정한다."""
    stdin_text("{not json")
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "JSON을 읽을 수 없습니다" in out
    assert "JSONDecodeError" not in out


def test_add_rejects_non_string_middle(stdin_text, ready, monkeypatch, capsys):
    """`middle` 이 문자열이 아니면 `str()` 로 뭉개지 않고 거부한다.

    이 테스트는 원래 `title` 을 대상으로 "잘못된 타입의 필드는 JSON 경로와
    함께 보고된다"를 고정했다. `tc add` 가 더 이상 `title` 을 읽지 않게 되면서
    `_one(title=...)` 은 그냥 무시되는 여분의 키가 되어 rc==0 으로 통과해
    버린다 — 코드가 실제로 읽는 필드로 과녁을 옮긴다. `{"middle": {"a": 1}}`
    은 truthy 라 필수 필드 검사를 통과하고, `str(item["middle"])` 이
    `{'a': 1}` 을 그대로 중분류 칸에 넣었다. steps/expected 와 같은 결함이다.
    """
    stdin_text(_one(middle={"a": 1}))
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "testcases[0].middle" in out
    assert "dict" in out
    assert _stored(ready) == []


def test_add_rejects_non_string_precondition(stdin_text, ready, monkeypatch, capsys):
    """선택 필드도 문자열이어야 한다 — 있으면 그대로 xlsx 칸에 들어간다."""
    stdin_text(_one(precondition=["a", "b"]))
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    assert "testcases[0].precondition" in capsys.readouterr().out
    assert _stored(ready) == []


def test_add_rejects_unknown_priority(stdin_text, ready, monkeypatch, capsys):
    """잘못된 priority 는 날 `ValueError` 대신 유효값을 알려주며 거부한다."""
    stdin_text(_one(priority="긴급"))
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "testcases[0].priority" in out
    assert "ValueError" not in out
    assert "High" in out                    # 유효값을 알려준다
    assert _stored(ready) == []


def test_add_names_the_offending_index(stdin_text, ready, monkeypatch, capsys):
    """여러 건 중 몇 번째가 틀렸는지 짚는다 — 앞의 두 건은 유효하다."""
    ok = {"middle": "m", "sub": "s", "steps": ["s"], "expected": ["e"]}
    payload = json.dumps({"testcases": [ok, ok, {**ok, "steps": "문자열"}]},
                         ensure_ascii=False)
    stdin_text(payload)
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "testcases[2].steps" in out
    assert "testcases[0]" not in out
    # 한 건이라도 틀리면 아무것도 저장하지 않는다
    assert _stored(ready) == []


def test_add_accepts_valid_priority_override(stdin_text, ready, monkeypatch):
    """검증이 정상 입력까지 막지 않는지 — 유효한 priority 는 그대로 쓰인다."""
    stdin_text(_one(priority="Low"))
    assert main(["tc", "add", "파티편성", "--family", "정상 경로",
                 "--origin", "interview", "--json", "-"]) == 0
    assert _stored(ready)[0].priority.value == "Low"


def test_add_sets_kind_and_priority_from_family(stdin_text, ready, monkeypatch):
    stdin_text(_payload())
    main(["tc", "add", "파티편성", "--family", "정상 경로",
          "--origin", "interview", "--json", "-"])
    with KnowledgeStore(ready / "starrail.db") as s:
        tc = s.testcases("파티편성")[0]
    assert tc.kind.value == "정상"
    assert tc.priority.value == "High"
    assert tc.origin.value == "인터뷰"
    assert tc.category_major == "파티편성"
    # `category_minor` 는 더 이상 계열의 대용품이 아니다 (커밋 8449bda) —
    # 계열은 `family` 필드가 실어 나른다. `category_minor` 는 이제 `middle`
    # (화면·메뉴 이름)을 담으므로, 계열이 제대로 전달됐는지는 `family` 로 본다.
    assert tc.family == "정상 경로"


def test_list_shows_unmet_slots(stdin_text, ready, monkeypatch, capsys):
    stdin_text(_payload())
    main(["tc", "add", "파티편성", "--family", "정상 경로",
          "--origin", "interview", "--json", "-"])
    assert main(["tc", "list", "파티편성"]) == 0
    out = capsys.readouterr().out
    assert "정상 동작" in out
    assert "재화 부족" in out  # 미충족 리포트


def test_add_rejects_unknown_status_slot_family(stdin_text, ready, monkeypatch, capsys):
    # cost 슬롯을 "모른다"로 답한 상태 — empty 와 다른 사유여야 한다.
    main(["slot", "set", "파티편성", "cost", "--status", "unknown"])
    capsys.readouterr()
    stdin_text(_payload())
    rc = main(["tc", "add", "파티편성", "--family", "재화 부족",
               "--origin", "inferred", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "cost" in out
    assert "사용자가 모른다고 답함" in out
    assert "슬롯이 비어 있음" not in out
    assert "해당 없음으로 표시됨" not in out


def test_add_rejects_na_status_slot_family(stdin_text, ready, monkeypatch, capsys):
    # cost 슬롯이 "해당 없음"으로 표시된 상태 — empty/unknown 과 다른 사유여야 한다.
    main(["slot", "set", "파티편성", "cost", "--status", "na"])
    capsys.readouterr()
    stdin_text(_payload())
    rc = main(["tc", "add", "파티편성", "--family", "재화 부족",
               "--origin", "inferred", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "cost" in out
    assert "해당 없음으로 표시됨" in out
    assert "슬롯이 비어 있음" not in out
    assert "사용자가 모른다고 답함" not in out


def test_add_reports_correct_added_and_kept_counts(stdin_text, ready, monkeypatch, capsys):
    payload = json.dumps({
        "testcases": [
            {
                "middle": "파티 편성", "sub": "정상 동작 1",
                "precondition": "파티 편성 화면",
                "steps": ["파티 적용을 누른다"],
                "expected": ["파티가 적용된다"],
                "rationale": "core_action 슬롯에서 도출",
            },
            {
                "middle": "파티 편성", "sub": "정상 동작 2",
                "precondition": "파티 편성 화면",
                "steps": ["다른 파티를 적용한다"],
                "expected": ["다른 파티가 적용된다"],
                "rationale": "core_action 슬롯에서 도출",
            },
        ]
    }, ensure_ascii=False)
    stdin_text(payload)
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 0
    out = capsys.readouterr().out
    # 페이로드의 testcases 수(2)와 정확히 일치해야 한다 — (added, kept) 언패킹
    # 순서가 뒤바뀌면 "TC 0건 저장 · 사람 손댄 2건 보존"처럼 거짓 실패로 보인다.
    assert "TC 2건 저장" in out
    assert "보존" not in out
    with KnowledgeStore(ready / "starrail.db") as s:
        assert len(s.testcases("파티편성")) == 2


def _titled(*titles):
    """소분류만 다른 유효한 TC 페이로드.

    이름은 `_titled` 로 남겨 뒀다 — 호출부가 여전히 "제목만 다른 TC 여러 개"를
    말하고 싶어 하고, 그 역할을 이제 `sub`(소분류)가 잇는다. `middle` 은
    구분에 쓰이지 않으므로 고정값이다.
    """
    return json.dumps({"testcases": [
        {"middle": "파티 편성", "sub": t, "steps": ["절차"], "expected": ["기대"]}
        for t in titles
    ]}, ensure_ascii=False)


def test_second_add_on_same_family_says_how_many_it_replaced(stdin_text, ready, capsys):
    """같은 계열에 `tc add` 를 두 번 부르면 **지운 개수를 소리 내어 말해야 한다** (BL2).

    실측 BEFORE — 격리된 스토어에서:

        1회차 → ✓ [정상 경로] TC 2건 저장   → tc list: A, B
        2회차 → ✓ [정상 경로] TC 1건 저장   → tc list: C 뿐  (A · B 증발, 무성)

    교체는 의도된 설계다 (`replace_generated` 가 존재하는 이유이고,
    `generated_hash` 가 사람 손댄 TC를 지켜준다). 결함은 파괴적 교체가 rc=0 ·
    성공 줄로, 무엇을 지웠는지 한 글자도 없이 끝난다는 것이다. 명령 이름은
    `add` 이고 SKILL.md 는 "계열마다 한 번씩" 만 말하므로, 사용자가 "이것도
    넣어주세요" 라고 해서 모델이 같은 계열을 한 번 더 부르는 순간 앞 배치가
    통째로 날아간다.
    """
    stdin_text(_titled("A", "B"))
    assert main(["tc", "add", "파티편성", "--family", "정상 경로",
                 "--origin", "interview", "--json", "-"]) == 0
    first = capsys.readouterr().out
    assert "TC 2건 저장" in first
    # 지운 게 없으면 `if deleted:` 블록이 통째로 나오지 않으므로 `⚠` 하나면
    # 충분하다. 여기 있던 `assert "교체" not in first` 는 지웠다 — "교체" 도
    # "다시 실행하세요" 도 모두 그 블록 안 문구라 새로 잡아내는 변이가 없는,
    # 이 브랜치가 아홉 번 만들어낸 "이웃의 부분문자열" 모양이었다.
    assert "⚠" not in first
    assert [t.category_sub for t in _stored(ready)] == ["A", "B"]

    stdin_text(_titled("C"))
    assert main(["tc", "add", "파티편성", "--family", "정상 경로",
                 "--origin", "interview", "--json", "-"]) == 0
    second = capsys.readouterr().out

    assert "TC 1건 저장" in second
    assert "⚠" in second                        # 눈에 띄어야 한다
    assert "기존 TC 2건" in second               # 몇 건을 지웠는지 정확히
    assert "다시 실행하세요" in second           # 다음 조치를 알린다
    assert [t.category_sub for t in _stored(ready)] == ["C"]


def test_replaced_count_excludes_preserved_user_testcases(stdin_text, ready, capsys):
    """`origin=user` 로 넣은 TC 는 보존된다 — 그건 지운 수에 세면 안 된다.

    보존과 삭제를 한 숫자로 뭉치면 "2건 지웠다" 가 거짓이 되고, 사용자는
    남아 있는 TC를 잃었다고 오해한다.
    """
    stdin_text(_titled("생성분"))
    main(["tc", "add", "파티편성", "--family", "정상 경로",
          "--origin", "interview", "--json", "-"])
    stdin_text(_titled("사람이 추가"))
    main(["tc", "add", "파티편성", "--family", "정상 경로",
          "--origin", "user", "--json", "-"])
    # 이 시점: '사람이 추가'(user) 만 남아 있다 — 앞의 생성분은 교체됐다
    assert [t.category_sub for t in _stored(ready)] == ["사람이 추가"]
    capsys.readouterr()          # 앞선 명령의 확인 문구를 버린다

    stdin_text(_titled("새 생성분"))
    assert main(["tc", "add", "파티편성", "--family", "정상 경로",
                 "--origin", "interview", "--json", "-"]) == 0
    out = capsys.readouterr().out

    assert "사람 손댄 1건 보존" in out
    assert "⚠" not in out                  # 지운 것이 없으므로 경고도 없다
    assert {t.category_sub for t in _stored(ready)} == {"사람이 추가", "새 생성분"}


def test_add_without_a_code_refuses_before_deleting_existing_rows(cfg_env, stdin_text, capsys, make_tc):
    """코드 없는 컨텐츠에 `tc add` 를 다시 부르면, 델리트가 시작되기 전에
    거절해야 한다 (Bug A) — 그리고 그 메시지는 `KeyError` 원문이 아니라 우리
    문구여야 한다.

    `ready` 픽스처는 항상 `--code PARTY` 로 만들어져 이 경로를 재현하지
    못한다 — 여기서는 코드 없이 직접 `slot init` 하고, 기존 행은 (코드가
    있어야 발급되는) 자동 id 대신 명시적 id 로 저장소에 직접 넣는다 —
    마스터 시절 DB에 이미 있던 행을 흉내낸 것이다.
    """
    main(["slot", "init", "파티편성", "--game", "starrail"])   # --code 없음
    main(["slot", "set", "파티편성", "core_action", "--status", "filled",
          "--value", "파티를 짜고 적용한다"])
    with KnowledgeStore(cfg_env / "starrail.db") as st:
        st.add_testcase("파티편성", "정상 경로",
                         make_tc(id="tc_old0", category_sub="케이스0"), ["core_action"])
        st.add_testcase("파티편성", "정상 경로",
                         make_tc(id="tc_old1", category_sub="케이스1"), ["core_action"])
    capsys.readouterr()

    stdin_text(_payload())
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "--code" in out
    assert "KeyError" not in out           # 날 예외가 아니라 우리 메시지여야 한다
    assert sorted(t.id for t in _stored(cfg_env)) == ["tc_old0", "tc_old1"]


def test_add_refuses_a_batch_with_duplicate_middle_sub_pairs(stdin_text, ready, capsys):
    """배치 안에서 (중분류, 소분류) 가 겹치면 거절한다 (Bug B) — 조용히
    번호를 물려주면 나중 것이 앞의 것을 지우는데 rc=0 · "2건 저장" 만 찍힌다.

    소분류 이름을 일부러 "중복"이 **아닌** 값으로 짓는다 — 예전 어서션
    (`assert "중복" in out`)은 소분류 이름 자체가 "중복"이라 이 가드의
    메시지가 실제로 찍혔는지와 무관하게 우연히 통과할 수 있었다(그 문자열은
    가드 메시지 안에서 소분류 값을 그대로 인용하는 자리에도 나온다). 가드가
    실제로 내는 고정 문구("이 배치에 두 번 있습니다")로 검사해야, 메시지가
    아예 안 나가도록 지워버리는 회귀를 이 테스트가 놓치지 않는다.
    """
    dup = json.dumps({"testcases": [
        {"middle": "파티 편성", "sub": "같은 케이스", "steps": ["s1"], "expected": ["e1"]},
        {"middle": "파티 편성", "sub": "같은 케이스", "steps": ["s2"], "expected": ["e2"]},
    ]}, ensure_ascii=False)
    stdin_text(dup)
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", "-"])
    assert rc == 1
    out = capsys.readouterr().out
    assert "이 배치에 두 번 있습니다" in out   # 가드의 실제 메시지 — 소분류 값과 무관한 고정 문구
    assert "ValueError" not in out         # 날 예외가 아니라 우리 메시지여야 한다
    assert _stored(ready) == []            # 아무것도 저장되지 않았다


def test_add_stores_inferred_origin(stdin_text, ready, monkeypatch):
    stdin_text(_payload())
    main(["tc", "add", "파티편성", "--family", "정상 경로",
          "--origin", "inferred", "--json", "-"])
    with KnowledgeStore(ready / "starrail.db") as s:
        tc = s.testcases("파티편성")[0]
    assert tc.origin.value == "추론됨"


def test_add_stores_user_origin(stdin_text, ready, monkeypatch):
    stdin_text(_payload())
    main(["tc", "add", "파티편성", "--family", "정상 경로",
          "--origin", "user", "--json", "-"])
    with KnowledgeStore(ready / "starrail.db") as s:
        tc = s.testcases("파티편성")[0]
    assert tc.origin.value == "사용자추가"




# --- 근거 철회 (I1) ------------------------------------------------------


def _withdraw_core_action():
    """`정상 경로` 의 근거였던 슬롯을 "해당 없음"으로 되돌린다."""
    main(["slot", "set", "파티편성", "core_action", "--status", "na"])


def test_list_marks_testcase_whose_evidence_was_withdrawn(stdin_text, ready, monkeypatch, capsys):
    """같은 출력이 "TC가 있다"와 "TC가 없다"를 동시에 말하면 안 된다.

    실측 BEFORE — `정상 경로` TC 를 만든 뒤 근거 슬롯을 NA 로 내리면:

        TC 1건 ... [정상 경로] 정상 동작  (인터뷰)          ← TC가 있다
        ⚠ ... core_action ... → 정상 경로 TC 없음  [해당 없음으로 표시됨]  ← 없다

    두 문장 중 하나는 반드시 거짓이고, 이 출력이 QA 담당자에게 가는 최종
    산출물이다. TC 는 **지우지 않는다** — 정정은 인터뷰의 정상 동작이고,
    사용자가 쌓아온 것을 도구가 조용히 버리면 안 된다.
    """
    stdin_text(_payload())
    main(["tc", "add", "파티편성", "--family", "정상 경로",
          "--origin", "interview", "--json", "-"])
    _withdraw_core_action()
    capsys.readouterr()          # 앞선 명령의 확인 문구를 버린다

    assert main(["tc", "list", "파티편성"]) == 0
    out = capsys.readouterr().out
    assert "정상 동작" in out              # TC 는 그대로 남아 있다
    assert "근거 철회됨" in out            # 그런데 근거가 없다고 표시된다
    assert "정상 경로 TC 없음" not in out  # 거짓말을 하지 않는다
    assert [t.category_sub for t in _stored(ready)] == ["정상 동작"]   # 삭제 안 함


def test_list_still_reports_families_that_really_have_no_tc(ready, capsys):
    """철회 표시가 진짜 미확인 항목 리포트를 삼키면 안 된다."""
    assert main(["tc", "list", "파티편성"]) == 0
    out = capsys.readouterr().out
    assert "재화 부족 TC 없음" in out
    assert "근거 철회됨" not in out


def test_list_stops_marking_once_evidence_is_restored(stdin_text, ready, monkeypatch, capsys):
    """슬롯을 다시 채우면 표시가 사라진다 — 표시가 영구 낙인이면 못 쓴다."""
    stdin_text(_payload())
    main(["tc", "add", "파티편성", "--family", "정상 경로",
          "--origin", "interview", "--json", "-"])
    _withdraw_core_action()
    main(["slot", "set", "파티편성", "core_action", "--status", "filled",
          "--value", "파티를 짜고 적용한다"])
    capsys.readouterr()          # 앞선 명령의 확인 문구를 버린다

    main(["tc", "list", "파티편성"])
    out = capsys.readouterr().out
    assert "근거 철회됨" not in out
    assert "정상 동작" in out


def test_export_xlsx_marks_withdrawn_evidence(stdin_text, ready, monkeypatch, capsys, tmp_path):
    """CLI 가 실제로 철회 정보를 익스포터까지 넘기는지 — 배선을 고정한다.

    `export_tc_excel` 이 아무리 잘 표시해도 `cmd_export` 가 안 넘기면 최종
    산출물은 그대로 자기모순이다.
    """
    from openpyxl import load_workbook

    stdin_text(_payload())
    main(["tc", "add", "파티편성", "--family", "정상 경로",
          "--origin", "interview", "--json", "-"])
    _withdraw_core_action()
    out_path = tmp_path / "out.xlsx"
    capsys.readouterr()          # 앞선 명령의 확인 문구를 버린다

    assert main(["export", "파티편성", "--out", str(out_path)]) == 0
    wb = load_workbook(out_path)
    assert wb.sheetnames == ["테스트케이스", "미확인 항목", "요약"]

    ws = wb["테스트케이스"]
    header = [c.value for c in ws[1]]
    assert "근거 상태" in header
    assert ws[2][header.index("근거 상태")].value == "근거 철회됨"

    skipped = wb["미확인 항목"]
    rows = [[c.value for c in r] for r in skipped.iter_rows(min_row=2)]
    withdrawn_row = next(r for r in rows if "core_action" in r)
    assert any(v and "근거 철회됨" in str(v) for v in withdrawn_row)


# --- tc plan 텍스트 모드 (T2 · M9) ---------------------------------------


def test_plan_text_lists_each_planned_family_with_kind_and_priority(ready, capsys):
    """`tc plan` 의 텍스트 모드가 통째로 미검증이었다 (M9a) — 모든 테스트가 `--json`.

    사람이 `qatc tc plan` 을 그냥 치면 보는 것이 이 화면이다. 계획 행을 아예
    출력하지 않아도 스위트가 초록이었다.
    """
    main(["slot", "set", "파티편성", "constraints", "--status", "filled",
          "--value", "최대 4명"])
    capsys.readouterr()          # 앞선 명령의 확인 문구를 버린다

    assert main(["tc", "plan", "파티편성"]) == 0
    out = capsys.readouterr().out

    assert "[파티편성] 생성 대상 계열 2개" in out
    planned_block = out.split("제외됨:", 1)[0]
    happy = next(ln for ln in planned_block.splitlines() if "정상 경로" in ln)
    assert "core_action" in happy          # 어느 슬롯이 근거인지
    assert "정상 / High" in happy          # TCKind / 기본 우선순위
    boundary = next(ln for ln in planned_block.splitlines() if "경계값" in ln)
    assert "constraints" in boundary
    assert "경계값 / Medium" in boundary


def test_plan_text_lists_skipped_families_with_their_reason(ready, capsys):
    """제외 블록도 미검증이었다 (M9b).

    이 블록이 없으면 사용자는 "왜 이 계열이 안 만들어지는가" 를 알 수 없고,
    무엇을 더 물어야 하는지도 알 수 없다 — 게이트의 판단이 보이지 않게 된다.
    """
    main(["slot", "set", "파티편성", "cost", "--status", "unknown"])
    main(["slot", "set", "파티편성", "failure", "--status", "na"])
    capsys.readouterr()          # 앞선 명령의 확인 문구를 버린다

    assert main(["tc", "plan", "파티편성"]) == 0
    out = capsys.readouterr().out
    assert "제외됨:" in out
    skipped_block = out.split("제외됨:", 1)[1]

    rows = {}
    for line in skipped_block.splitlines():
        for family in ("재화 부족", "실패 경로", "진입 경로"):
            if family in line:
                rows[family] = line
    assert set(rows) == {"재화 부족", "실패 경로", "진입 경로"}

    # 상태 세 가지가 서로 다른 사유로 보여야 한다 (4상태 설계의 요점)
    assert "cost" in rows["재화 부족"] and "사용자가 모른다고 답함" in rows["재화 부족"]
    assert "failure" in rows["실패 경로"] and "해당 없음으로 표시됨" in rows["실패 경로"]
    assert "entry" in rows["진입 경로"] and "슬롯이 비어 있음" in rows["진입 경로"]

    # 근거가 있는 계열은 제외 블록에 없다
    assert "정상 경로" not in skipped_block


# --- tc add --json <파일 경로> (T2 · M11) --------------------------------


def test_add_reads_the_payload_from_a_file_path(ready, tmp_path, capsys):
    """`--json <파일 경로>` 갈래가 통째로 미검증이었다 (M11a).

    SKILL.md 는 히어독(`--json -`)을 쓰지만 도움말은 "JSON 파일 경로 또는 '-'"
    라고 약속한다. 표준입력만 읽도록 되돌려도 스위트가 초록이었다.
    """
    path = tmp_path / "tc.json"
    path.write_text(_payload("파일에서 읽은 TC"), encoding="utf-8")

    assert main(["tc", "add", "파티편성", "--family", "정상 경로",
                 "--origin", "interview", "--json", str(path)]) == 0
    assert [t.category_sub for t in _stored(ready)] == ["파일에서 읽은 TC"]


# --- tc add 입력 계약: middle/sub, 경고 두 종 (T4) -----------------------


def test_middle_and_sub_are_stored(ready, stdin_text, capsys):
    stdin_text(_payload(middle="신규 계정 연동", sub="비밀번호 불일치"))
    assert main(["tc", "add", "파티편성", "--family", "정상 경로",
                 "--origin", "inferred", "--json", "-"]) == 0
    with KnowledgeStore(ready / "starrail.db") as st:
        tc = st.testcases("파티편성")[0]
    assert (tc.category_minor, tc.category_sub) == ("신규 계정 연동", "비밀번호 불일치")


@pytest.mark.parametrize("field", ["middle", "sub"])
def test_a_blank_middle_or_sub_is_refused(ready, stdin_text, capsys, field):
    """제로폭 공백은 `strip()` 이 못 지운다 — `is_blank` 로 판정한다."""
    stdin_text(_payload(**{field: "  \u200b "}))
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "inferred", "--json", "-"])
    assert rc == 1
    assert "비어" in capsys.readouterr().out


def test_an_empty_expected_is_refused(ready, stdin_text, capsys):
    """확인할 것이 없는 TC 는 TC 가 아니다."""
    stdin_text(_payload(expected=[]))
    assert main(["tc", "add", "파티편성", "--family", "정상 경로",
                 "--origin", "inferred", "--json", "-"]) == 1


def test_more_than_six_expected_warns_but_saves(ready, stdin_text, capsys):
    """규칙 3은 기계적이라 셀 수 있다. 다만 **막지는 않는다** — 나눌지는
    판단이고, 막으면 그 판단을 표현할 방법이 없어진다."""
    stdin_text(_payload(expected=[f"확인 {n}" for n in range(7)]))
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "inferred", "--json", "-"])
    assert rc == 0, "경고여야 하는데 막았습니다"
    out = capsys.readouterr().out
    assert "6개" in out and "나누" in out
    with KnowledgeStore(ready / "starrail.db") as st:
        assert len(st.testcases("파티편성")) == 1


def test_a_long_sub_warns_but_saves(ready, stdin_text, capsys):
    """길다는 것은 결과를 밀어 넣었다는 신호이지 그 자체가 오류는 아니다."""
    stdin_text(_payload(sub="비밀번호 두 필드가 불일치하면 연동하기 버튼이 비활성으로 유지된다"))
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "inferred", "--json", "-"])
    assert rc == 0
    assert "소분류" in capsys.readouterr().out


def test_a_judgement_rule_is_never_enforced_by_code(ready, stdin_text, capsys):
    """기대결과가 2개라는 이유만으로 막으면 안 된다.

    한때 `expected` 를 정확히 1개로 강제하려 했다가 철회했다 — 회원가입의
    DB 저장과 이메일 발송처럼 한 문장으로 이어 쓸 수 없는 독립 결과가 있고,
    반대로 화면 전환과 문구 노출처럼 한 세트인 것도 있다. 그 구분은 판단이라
    코드가 셀 수 없다. 누가 선의로 하드 거부를 되살리면 이 테스트가 막는다.
    """
    stdin_text(_payload(expected=["메인 페이지로 이동한다", "환영 문구가 노출된다"]))
    assert main(["tc", "add", "파티편성", "--family", "정상 경로",
                 "--origin", "inferred", "--json", "-"]) == 0


def test_add_on_missing_file_says_it_could_not_read_the_json(ready, tmp_path, capsys):
    """없는 파일을 주면 날 `FileNotFoundError` 대신 우리 메시지가 나와야 한다 (M11b).

    `except (OSError, json.JSONDecodeError)` 에서 `OSError` 를 빼도 초록이었다 —
    그러면 `cli.py` 의 범용 핸들러가 `오류: FileNotFoundError: [Errno 2] ...` 를
    앞줄 공백과 함께 뱉는다. 인터뷰를 진행하는 모델에게 파이썬 예외 이름은
    다음에 무엇을 할지 알려주지 않는다.
    """
    missing = tmp_path / "없는파일.json"
    rc = main(["tc", "add", "파티편성", "--family", "정상 경로",
               "--origin", "interview", "--json", str(missing)])
    assert rc == 1
    out = capsys.readouterr().out
    assert "JSON을 읽을 수 없습니다" in out
    assert "FileNotFoundError" not in out
    assert _stored(ready) == []
