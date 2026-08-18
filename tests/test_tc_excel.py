import os
import stat

import pytest
from openpyxl import load_workbook

from qatc.export.tc_excel import ExportBlocked, clean_cell, export_tc_excel
from qatc.knowledge.gate import FamilySkip
from qatc.knowledge.models import SlotStatus
from qatc.models import TCOrigin


def test_clean_cell_strips_control_characters():
    assert clean_cell("A\x03B") == "AB"
    assert clean_cell("정상\x00문자") == "정상문자"


def test_clean_cell_keeps_newline_and_tab():
    assert clean_cell("가\n나\t다") == "가\n나\t다"


def test_clean_cell_passes_through_normal_text():
    assert clean_cell("파티 적용") == "파티 적용"


def test_export_creates_three_sheets(make_tc, tmp_path):
    p = export_tc_excel("파티편성", [make_tc()], [], tmp_path / "out.xlsx")
    wb = load_workbook(p)
    assert wb.sheetnames == ["테스트케이스", "미확인 항목", "요약"]


def test_export_writes_testcase_row(make_tc, tmp_path):
    # `title` 은 더 이상 행에 안 쓰인다 (task-2) — 행을 식별하는 값은
    # `category_sub` 로 옮겨졌다. 확인하려는 것은 그대로다: 지정한 값이
    # 실제로 행에 쓰이고, 출처도 함께 나온다는 것.
    p = export_tc_excel("파티편성", [make_tc(category_sub="정상 동작")], [], tmp_path / "out.xlsx")
    ws = load_workbook(p)["테스트케이스"]
    values = [c.value for c in ws[2]]
    assert "정상 동작" in values
    assert "인터뷰" in values


def test_export_survives_control_characters(make_tc, tmp_path):
    # 이 케이스가 예전 export 를 통째로 죽였다 (IllegalCharacterError).
    # 제어문자를 실어 나르던 필드가 `title` 에서 `category_sub` 로 바뀌었을 뿐,
    # "행에 쓰이는 문자열은 clean_cell 을 반드시 거친다" 는 확인 대상은 같다.
    bad = make_tc(category_sub="제어문자\x03포함")
    p = export_tc_excel("파티편성", [bad], [], tmp_path / "out.xlsx")
    ws = load_workbook(p)["테스트케이스"]
    assert any(c.value == "제어문자포함" for c in ws[2])


def test_export_lists_skipped_families(make_tc, tmp_path):
    skip = FamilySkip("재화 부족", "cost", "무엇을 소모하는가", SlotStatus.EMPTY)
    p = export_tc_excel("파티편성", [make_tc()], [skip], tmp_path / "out.xlsx")
    ws = load_workbook(p)["미확인 항목"]
    row = [c.value for c in ws[2]]
    assert "cost" in row
    assert "재화 부족" in row


def test_export_survives_control_characters_in_skipped_sheet(make_tc, tmp_path):
    # clean_cell 이 _sheet_skipped 에서 빠지면 이 케이스가 IllegalCharacterError 로 죽는다 —
    # prompt_hint 는 인터뷰 중 사람이 받아적거나 OCR 로 채워질 수 있는 자유 텍스트다.
    skip = FamilySkip("재화 부족", "cost", "무엇을\x03소모하는가", SlotStatus.EMPTY)
    p = export_tc_excel("파티편성", [make_tc()], [skip], tmp_path / "out.xlsx")
    ws = load_workbook(p)["미확인 항목"]
    row = [c.value for c in ws[2]]
    assert "무엇을소모하는가" in row
    assert not any(v is not None and "\x03" in str(v) for v in row)


def _summary_value(ws, label: str) -> str:
    for row in ws.iter_rows(min_row=2):
        if row[0].value == label:
            return row[1].value
    raise AssertionError(f"'{label}' 행을 요약 시트에서 찾을 수 없습니다")


def test_export_summary_counts_by_origin(make_tc, tmp_path):
    cases = [make_tc(origin=TCOrigin.INTERVIEW), make_tc(origin=TCOrigin.INFERRED)]
    p = export_tc_excel("파티편성", cases, [], tmp_path / "out.xlsx")
    ws = load_workbook(p)["요약"]
    # "읽는 방법" 설명 칸에 '인터뷰'·'추론됨' 문구가 항상 들어있어 시트 전체 텍스트를
    # 뒤지면 데이터가 비어 있어도 통과한다. 라벨로 실제 행을 찾아 집계값을 확인한다.
    assert _summary_value(ws, "출처별") == "인터뷰 1, 추론됨 1"
    assert _summary_value(ws, "유형별") == "정상 2"


def test_locked_target_raises_export_blocked_with_next_action(tmp_path):
    """Excel 이 열어둔 파일에 쓰려 할 때. 읽기 전용으로 같은 조건을 만든다."""
    out = tmp_path / "잠김.xlsx"
    export_tc_excel("컨텐츠", [], [], out)          # 1차 — 정상
    os.chmod(out, stat.S_IREAD)
    try:
        with pytest.raises(ExportBlocked) as e:
            export_tc_excel("컨텐츠", [], [], out)  # 2차 — 막힘
        msg = str(e.value)
        assert str(out) in msg                       # 어느 파일인지
        assert "Excel" in msg and "닫" in msg        # 다음 조치
        assert "PermissionError" not in msg          # 예외 이름을 노출하지 않는다
        # 원인 체인이 보존돼야 원인 규명이 가능하다 (`from exc`) — 타입까지
        # 확인해야 `from None` 이나 엉뚱한 예외로 체이닝해도 잡아낸다.
        assert isinstance(e.value.__cause__, PermissionError)
    finally:
        os.chmod(out, stat.S_IWRITE)                 # 다른 테스트를 위해 되돌린다


def test_export_blocked_is_an_oserror():
    """호출자가 OSError 로도 잡을 수 있어야 한다."""
    assert issubclass(ExportBlocked, OSError)


def test_export_survives_control_characters_in_summary_sheet(make_tc, tmp_path):
    # clean_cell 이 _sheet_summary 에서 빠지면 이 케이스가 IllegalCharacterError 로 죽는다.
    p = export_tc_excel("파티\x03편성", [make_tc()], [], tmp_path / "out.xlsx")
    ws = load_workbook(p)["요약"]
    values = [c.value for row in ws.iter_rows() for c in row]
    assert "파티편성" in values
    assert not any(v is not None and "\x03" in str(v) for v in values)


def test_export_with_no_testcases_still_writes(tmp_path):
    p = export_tc_excel("파티편성", [], [], tmp_path / "out.xlsx")
    assert p.exists()


# --- 근거 철회 (I1) ------------------------------------------------------


def test_testcase_sheet_marks_withdrawn_family(make_tc, tmp_path):
    p = export_tc_excel("파티편성", [make_tc()], [], tmp_path / "out.xlsx",
                        withdrawn={"정상 경로"})
    ws = load_workbook(p)["테스트케이스"]
    header = [c.value for c in ws[1]]
    assert "근거 상태" in header
    assert ws[2][header.index("근거 상태")].value == "근거 철회됨"


def test_testcase_sheet_marks_live_family_as_valid(make_tc, tmp_path):
    """철회가 없으면 같은 칸이 "유효" 여야 한다 — 빈 칸은 "누락"으로 읽힌다."""
    p = export_tc_excel("파티편성", [make_tc()], [], tmp_path / "out.xlsx")
    ws = load_workbook(p)["테스트케이스"]
    header = [c.value for c in ws[1]]
    assert ws[2][header.index("근거 상태")].value == "유효"


def test_skipped_sheet_does_not_claim_a_family_has_no_tc_when_it_does(make_tc, tmp_path):
    """"만들지 못한 계열" 은 TC 가 실제로 있는 계열에 대해 거짓말이다.

    실측 BEFORE — `테스트케이스` 시트 2행에 `정상 경로` TC 가 있는데
    `미확인 항목` 시트가 같은 계열을 "만들지 못한 계열" 로 올렸다.
    """
    skip = FamilySkip("정상 경로", "core_action", "주 동작은 무엇인가", SlotStatus.NA)
    p = export_tc_excel("파티편성", [make_tc()], [skip], tmp_path / "out.xlsx",
                        withdrawn={"정상 경로"})
    ws = load_workbook(p)["미확인 항목"]
    assert "만들지 못한 계열" not in [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    assert any(v and "근거 철회됨" in str(v) for v in row)
    assert not any(v and "TC 없음" in str(v) for v in row)


def test_skipped_sheet_still_says_no_tc_when_there_really_is_none(tmp_path):
    skip = FamilySkip("재화 부족", "cost", "무엇을 소모하는가", SlotStatus.EMPTY)
    p = export_tc_excel("파티편성", [], [skip], tmp_path / "out.xlsx")
    ws = load_workbook(p)["미확인 항목"]
    row = [c.value for c in ws[2]]
    assert any(v and "TC 없음" in str(v) for v in row)


def test_summary_counts_withdrawn_testcases(make_tc, tmp_path):
    cases = [make_tc(title="철회된 것"), make_tc(title="살아 있는 것")]
    cases[1].category_minor = "경계값"
    cases[1].family = "경계값"          # 철회 판정은 family 를 보므로 함께 바꾼다
    p = export_tc_excel("파티편성", cases, [], tmp_path / "out.xlsx",
                        withdrawn={"정상 경로"})
    ws = load_workbook(p)["요약"]
    assert _summary_value(ws, "근거 철회된 TC") == "1"


def test_the_sheet_has_the_three_level_hierarchy_and_no_title(tmp_path, make_tc):
    """대·중·소가 표에 있고 `제목`·`유형` 칸은 없어야 한다.

    헤더를 통째로 비교한다 — 부분 문자열로 보면 컬럼 순서가 뒤바뀌거나
    하나가 사라져도 통과한다.
    """
    from openpyxl import load_workbook

    tc = make_tc(title="쓰이지 않는 옛 제목")
    tc.category_major, tc.category_minor = "로그인", "신규 계정 연동"
    tc.category_sub, tc.family = "비밀번호 불일치", "경계값"
    out = export_tc_excel("로그인", [tc], [], tmp_path / "o.xlsx", set())
    ws = load_workbook(out)["테스트케이스"]

    header = [c.value for c in ws[1]]
    assert header == ["TC ID", "대분류", "중분류", "소분류", "사전조건", "절차",
                      "기대결과", "우선순위", "출처", "근거", "근거 상태"]
    assert "제목" not in header
    assert "유형" not in header


def test_the_row_carries_the_hierarchy_in_order(tmp_path, make_tc):
    from openpyxl import load_workbook

    tc = make_tc(title="쓰이지 않는 옛 제목")
    tc.category_major, tc.category_minor = "로그인", "신규 계정 연동"
    tc.category_sub, tc.family = "비밀번호 불일치", "경계값"
    out = export_tc_excel("로그인", [tc], [], tmp_path / "o.xlsx", set())
    ws = load_workbook(out)["테스트케이스"]

    row = [c.value for c in ws[2]]
    assert row[1:4] == ["로그인", "신규 계정 연동", "비밀번호 불일치"]
    assert "쓰이지 않는 옛 제목" not in row
    # M7: _ORIGIN_FILL 이 칠하는 열이 실제 출처(9열)인지 확인한다. 값은
    # 색칠 위치와 무관하게 항상 같으므로 값 비교만으로는 옛 10열에 색칠이
    # 남아도 통과해버린다 — 9열이 정말 칠해졌는지(그리고 10열은 비어
    # 있는지) fill 자체를 봐야 잡힌다.
    assert ws.cell(row=2, column=9).value == tc.origin.value
    assert ws.cell(row=2, column=9).fill.fill_type == "solid"
    assert ws.cell(row=2, column=10).fill.fill_type is None


def test_withdrawal_still_shows_on_the_row(tmp_path, make_tc):
    """계열 컬럼이 사라져도 근거 철회 표시는 행에 남는다 — 읽는 사람이
    이 행을 믿어도 되는지 판단하는 유일한 단서다."""
    from openpyxl import load_workbook

    tc = make_tc()
    tc.family, tc.category_sub = "경계값", "비밀번호 불일치"
    out = export_tc_excel("로그인", [tc], [], tmp_path / "o.xlsx", {"경계값"})
    ws = load_workbook(out)["테스트케이스"]
    assert "철회" in str(ws.cell(row=2, column=11).value)


def test_locked_output_directory_also_raises_export_blocked(monkeypatch, tmp_path):
    """출력 **디렉터리** 를 만들 수 없을 때도 같은 안내가 나와야 한다.

    `wb.save` 는 Task 5 에서 막혔지만 `path.parent.mkdir` 이 try 밖에 있었다.

    실측 (이 환경): `os.chmod(dir, stat.S_IREAD)` 로 디렉터리를 잠그고 그 밑에
    `mkdir` 을 시도해도 Windows 에서는 자식 생성이 막히지 않았다 —
    `ExportBlocked` 는커녕 어떤 예외도 나지 않고 그대로 성공했다 (브리핑의 경고와
    일치). 그래서 OS 잠금 대신 `Path.mkdir` 자체를 실패하게 만들어, `mkdir` 이
    `wb.save` 와 같은 `try`/`except PermissionError` 안에 있는지를 직접 검증한다.
    """
    from pathlib import Path

    import pytest

    from qatc.export.tc_excel import ExportBlocked, export_tc_excel

    def _denied(self, *a, **k):
        raise PermissionError(13, "Permission denied", str(self))

    monkeypatch.setattr(Path, "mkdir", _denied)
    with pytest.raises(ExportBlocked) as e:
        export_tc_excel("컨텐츠", [], [], tmp_path / "하위" / "out.xlsx")
    msg = str(e.value)
    assert "PermissionError" not in msg
    # 브리핑: "메시지는 wb.save 쪽과 같은 문장을 쓴다 — 사용자에게는 같은 상황이다."
    # mkdir 실패도 wb.save 실패(test_locked_target_raises_export_blocked_with_next_action)와
    # 같은 안내를 담아야 한다 — 별도 메시지로 갈라지는 것을 여기서 막는다.
    assert "Excel" in msg and "닫" in msg
