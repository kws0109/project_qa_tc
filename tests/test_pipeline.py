"""저장소·플로우·익스포트·비용 통합 테스트."""

from __future__ import annotations

import numpy as np
import pytest

from qatc.analyze.cluster import ClusterResult, FrameFeatures
from qatc.analyze.flow import build_flow, describe_action
from qatc.analyze.hashing import ScreenSignature
from qatc.llm.cost import CostTracker, Usage, model_rates
from qatc.models import (
    CaptureReason,
    ElementKind,
    Frame,
    FlowGraph,
    InputEvent,
    InputKind,
    NormRect,
    ScreenState,
    SessionMeta,
    TCKind,
    TCOrigin,
    TestCase,
    Transition,
    UIElement,
)
from qatc.profiles import GameProfile, generic_profile
from qatc.storage import SessionStore, utcnow


# ---------------------------------------------------------------- 저장소


@pytest.fixture()
def store(tmp_path):
    meta = SessionMeta(
        id="t_sess", profile_name="generic", game_name="테스트",
        started_at=utcnow(), client_w=1280, client_h=720,
    )
    s = SessionStore.create(tmp_path, meta)
    yield s
    s.close()


def test_store_roundtrip_frames_and_events(store):
    store.add_frames([
        Frame(id=f"f{i}", session_id="t_sess", ts=i * 0.5, path=f"frames/f{i}.jpg",
              reason=CaptureReason.POST_SETTLED, client_w=1280, client_h=720, event_id="e1")
        for i in range(3)
    ])
    store.add_event(InputEvent(id="e1", session_id="t_sess", ts=1.0,
                               kind=InputKind.CLICK, nx=0.5, ny=0.5))
    assert len(store.frames()) == 3
    assert len(store.frames_for_event("e1")) == 3
    assert store.event("e1").kind is InputKind.CLICK
    assert store.frame("f1").ts == 0.5


def test_store_graph_roundtrip_preserves_user_layer(store):
    graph = FlowGraph(session_id="t_sess")
    state = ScreenState(id="st_000")
    state.user.name = "확정 화면"
    state.user.locked = True
    graph.states["st_000"] = state
    store.save_graph(graph)

    restored = store.load_graph()
    assert restored.states["st_000"].user.name == "확정 화면"
    assert restored.states["st_000"].user.locked
    assert (store.root / "flow.json").exists()


def test_store_graph_save_is_atomic_replacement(store):
    """상태 병합으로 노드가 줄었을 때 옛 노드가 남으면 안 된다."""
    graph = FlowGraph(session_id="t_sess")
    graph.states["a"] = ScreenState(id="a")
    graph.states["b"] = ScreenState(id="b")
    store.save_graph(graph)
    del graph.states["b"]
    store.save_graph(graph)
    assert set(store.load_graph().states) == {"a"}


def test_store_ocr_cache(store):
    assert store.ocr_get("h1") is None
    store.ocr_put("h1", [{"text": "캐릭터", "rect": [0, 0, 1, 1], "confidence": 0.9}])
    assert store.ocr_get("h1")[0]["text"] == "캐릭터"


def test_store_testcases_roundtrip(store):
    store.save_testcases([
        TestCase(id="TC-1", title="제목", steps=["a"], expected=["b"],
                 kind=TCKind.EXCEPTION, origin=TCOrigin.INFERRED)
    ])
    back = store.testcases()
    assert len(back) == 1 and back[0].origin is TCOrigin.INFERRED


def test_store_llm_ledger(store):
    store.log_llm_call("c1", "명명", "claude-sonnet-5", 1000, 500, cost_usd=0.01)
    totals = store.llm_totals()
    assert totals["calls"] == 1 and totals["input_tokens"] == 1000


def test_open_missing_session_raises(tmp_path):
    with pytest.raises(FileNotFoundError):
        SessionStore.open_existing(tmp_path / "없음")


# ---------------------------------------------------------------- 플로우


def _frame(fid: str, ts: float, reason: CaptureReason, event_id: str | None) -> Frame:
    return Frame(id=fid, session_id="s", ts=ts, path=f"frames/{fid}.jpg",
                 reason=reason, client_w=1280, client_h=720, event_id=event_id)


def _feature(fid: str, ts: float, settled: bool = True) -> FrameFeatures:
    img = np.full((64, 64, 3), 40, np.uint8)
    return FrameFeatures(frame_id=fid, ts=ts, sig=ScreenSignature.of(img), is_settled=settled)


def test_build_flow_creates_transition_from_burst():
    """PRE_ACTION 화면 → 입력 → POST_SETTLED 화면 이 하나의 전이가 된다."""
    frames = [
        _frame("pre", 0.9, CaptureReason.PRE_ACTION, "e1"),
        _frame("fast", 1.25, CaptureReason.POST_FAST, "e1"),
        _frame("settled", 2.5, CaptureReason.POST_SETTLED, "e1"),
    ]
    features = [_feature("pre", 0.9), _feature("settled", 2.5)]
    labels = np.array([0, 1])
    result = ClusterResult(labels=labels, clusters={0: [0], 1: [1]}, similarity=np.eye(2))
    event = InputEvent(id="e1", session_id="s", ts=1.0, kind=InputKind.CLICK, nx=0.3, ny=0.2)

    built = build_flow("s", frames, [event], features, result)
    assert len(built.graph.transitions) == 1
    t = built.graph.transitions[0]
    assert t.from_state != t.to_state
    assert built.no_change_events == 0


def test_build_flow_marks_no_change_as_self_loop():
    """클릭했는데 화면이 안 바뀐 것도 의미 있는 관측이다."""
    frames = [
        _frame("pre", 0.9, CaptureReason.PRE_ACTION, "e1"),
        _frame("settled", 2.5, CaptureReason.POST_SETTLED, "e1"),
    ]
    features = [_feature("pre", 0.9), _feature("settled", 2.5)]
    result = ClusterResult(labels=np.array([0, 0]), clusters={0: [0, 1]}, similarity=np.eye(2))
    event = InputEvent(id="e1", session_id="s", ts=1.0, kind=InputKind.CLICK, nx=0.3, ny=0.2)

    built = build_flow("s", frames, [event], features, result)
    assert built.no_change_events == 1
    assert built.graph.transitions[0].is_self_loop


def test_build_flow_counts_orphan_events():
    event = InputEvent(id="e9", session_id="s", ts=1.0, kind=InputKind.CLICK, nx=0.1, ny=0.1)
    result = ClusterResult(labels=np.array([0]), clusters={0: [0]}, similarity=np.eye(1))
    built = build_flow("s", [_frame("f", 0.5, CaptureReason.IDLE_CHANGE, None)],
                       [event], [_feature("f", 0.5)], result)
    assert built.orphan_events == 1


def test_describe_action_prefers_element_label():
    """TC 문구 품질의 핵심 — 좌표 대신 버튼 이름을 쓴다."""
    event = InputEvent(id="e", session_id="s", ts=0.0, kind=InputKind.CLICK, nx=0.4, ny=0.3)
    element = UIElement(NormRect(0.35, 0.25, 0.1, 0.1), ElementKind.BUTTON, label="강화하기")
    assert describe_action(event, element) == "[강화하기] 클릭"
    assert "0.40" in describe_action(event, None)


def test_describe_action_uses_profile_key_hints():
    profile = GameProfile(key="g", name="G", key_hints={"esc": "뒤로가기"})
    event = InputEvent(id="e", session_id="s", ts=0.0, kind=InputKind.KEY, key="esc")
    assert "뒤로가기" in describe_action(event, None, profile)


# ---------------------------------------------------------------- 프로파일


def test_profile_matches_by_process_over_title():
    profile = GameProfile(key="g", name="G", title_regex="^정확한제목$", process_name="Game.exe")
    assert profile.matches_window("전혀 다른 제목", "Game.exe")
    assert profile.matches_window("정확한제목", "")
    assert not profile.matches_window("다른 제목", "Other.exe")


def test_generic_profile_matches_nothing():
    """폴백 프로파일은 아무 창이나 잡으면 안 된다 — 사용자가 직접 골라야 한다."""
    assert not generic_profile().matches_window("아무 창", "any.exe")


def test_profile_loading(tmp_path):
    from qatc.profiles import load_profiles

    (tmp_path / "test.yaml").write_text(
        "name: 테스트게임\nwindow:\n  process: Test.exe\n"
        "capture_roi: [0.0, 0.05, 1.0, 0.95]\n"
        "static_ignore:\n  - {x: 0.8, y: 0.0, w: 0.2, h: 0.1, why: 시계}\n",
        encoding="utf-8",
    )
    (tmp_path / "broken.yaml").write_text("name: [불완전\n", encoding="utf-8")

    profiles = load_profiles(tmp_path)
    assert "test" in profiles
    assert "broken" not in profiles          # 깨진 파일은 건너뛰되 전체를 막지 않는다
    assert profiles["test"].capture_roi.y == pytest.approx(0.05)
    assert len(profiles["test"].ignore_rects) == 1


# ---------------------------------------------------------------- 비용


def test_sonnet_intro_pricing_expires():
    from datetime import date

    assert model_rates("claude-sonnet-5", date(2026, 8, 13)) == (2.00, 10.00)
    assert model_rates("claude-sonnet-5", date(2026, 9, 1)) == (3.00, 15.00)


def test_unknown_model_uses_conservative_rate():
    """모르는 모델은 비싸게 잡는다 — 예산을 넘겨 놀라게 하는 것보다 낫다."""
    assert model_rates("claude-미래모델") == (5.00, 25.00)


def test_cache_read_is_cheaper_than_fresh_input():
    fresh = Usage(input_tokens=10_000).cost_usd("claude-opus-5")
    cached = Usage(cache_read=10_000).cost_usd("claude-opus-5")
    assert cached < fresh / 5


def test_budget_gate():
    tracker = CostTracker(budget_usd=0.01)
    assert not tracker.over_budget
    tracker.record("claude-opus-5", "TC 생성", Usage(input_tokens=100_000, output_tokens=50_000))
    assert tracker.over_budget
    assert tracker.remaining_usd == 0.0


def test_usage_addition():
    total = Usage(input_tokens=10, output_tokens=5) + Usage(input_tokens=3, cache_read=7)
    assert (total.input_tokens, total.output_tokens, total.cache_read) == (13, 5, 7)


# ---------------------------------------------------------------- 익스포트


def _sample_graph() -> FlowGraph:
    graph = FlowGraph(session_id="t_sess")
    for sid, name in (("st_000", "홈"), ("st_001", "캐릭터")):
        state = ScreenState(id=sid)
        state.user.name = name
        graph.states[sid] = state
    t = Transition(id="tr_1", from_state="st_000", to_state="st_001", action_desc="[캐릭터] 클릭")
    graph.transitions.append(t)
    graph.step_order.append(t.id)
    return graph


def test_mermaid_marks_uncovered_edges():
    from qatc.export.mermaid import render_mermaid

    def edge_lines(text: str) -> list[str]:
        # 범례 주석(%%)에도 ⚠가 들어가므로 간선 줄만 본다
        return [l for l in text.splitlines() if "-->" in l]

    graph = _sample_graph()
    uncovered = render_mermaid(graph, [])
    assert any("⚠" in l for l in edge_lines(uncovered))
    assert "stroke-dasharray" in uncovered

    covered = render_mermaid(graph, [TestCase(id="t", edge_ids=["tr_1"])])
    assert not any("⚠" in l for l in edge_lines(covered))
    assert "stroke-dasharray" not in covered


def test_mermaid_escapes_unsafe_labels():
    from qatc.export.mermaid import render_mermaid

    graph = _sample_graph()
    graph.states["st_000"].user.name = 'A "인용" \n 개행 --> 화살표'
    body = render_mermaid(graph, [])
    label_line = next(l for l in body.splitlines() if l.strip().startswith("st_000 :"))
    assert '"' not in label_line and "\n" not in label_line.strip()


def test_excel_export_has_all_sheets(store, tmp_path):
    from openpyxl import load_workbook

    from qatc.export.excel import export_excel

    graph = _sample_graph()
    store.save_graph(graph)
    cases = [
        TestCase(id="TC-1", category_major="캐릭터", title="정상", steps=["a"], expected=["b"],
                 origin=TCOrigin.RECORDED, edge_ids=["tr_1"]),
        TestCase(id="TC-2", category_major="캐릭터", title="추론", steps=["a"], expected=["b"],
                 origin=TCOrigin.INFERRED),
    ]
    path = export_excel(store, graph, cases, tmp_path / "out.xlsx", embed_thumbnails=False)
    wb = load_workbook(path)
    assert wb.sheetnames == ["테스트케이스", "커버리지", "요약"]
    ws = wb["테스트케이스"]
    assert ws.max_row == 3
    # 출처별로 배경색이 달라야 한다 (추론됨이 눈에 띄어야 함)
    assert ws.cell(2, 10).fill.fgColor.rgb != ws.cell(3, 10).fill.fgColor.rgb


def test_excel_export_with_no_testcases(store, tmp_path):
    """TC가 없어도 커버리지·요약 시트는 나와야 한다 — LLM 없이 쓰는 경로."""
    from qatc.export.excel import export_excel

    graph = _sample_graph()
    store.save_graph(graph)
    path = export_excel(store, graph, [], tmp_path / "empty.xlsx", embed_thumbnails=False)
    assert path.exists()
