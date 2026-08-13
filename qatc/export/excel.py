"""Excel 테스트케이스 시트 출력 (openpyxl).

국내 게임사 QA 실무 포맷을 따른다. 시트 세 장으로 나눈다.

* **테스트케이스** — 실제 TC 목록. 근거 스크린샷 썸네일과 원본 링크가 붙는다.
* **커버리지** — 관측된 전이 중 어떤 TC도 커버하지 않은 것. **가장 실무적인 시트**다.
  "무엇을 테스트하지 않았는가"가 QA에서 가장 값진 정보이기 때문이다.
* **요약** — 분류·유형·출처별 집계와 세션 메타데이터.

**썸네일은 삽입하고 원본은 링크한다.** 전체 이미지를 넣으면 TC 200건짜리 파일이
수백 MB가 되어 열리지 않는다. 120px 썸네일이면 리뷰어가 "이 화면 맞나" 확인하기에
충분하고, 자세히 볼 때는 링크로 원본을 연다.
"""

from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Iterable, Sequence

import cv2
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..models import FlowGraph, SessionMeta, TCOrigin, TestCase, coverage
from ..storage import SessionStore

#: 썸네일 가로 픽셀. 행 높이가 여기 맞춰 결정된다.
THUMB_W = 150

HEADERS = [
    ("TC ID", 13),
    ("대분류", 12),
    ("중분류", 14),
    ("테스트 항목", 34),
    ("사전조건", 26),
    ("테스트 절차", 46),
    ("기대 결과", 46),
    ("우선순위", 9),
    ("유형", 9),
    ("출처", 10),
    ("근거", 22),
    ("작성 근거", 30),
]

#: 출처별 배경색. **추론됨을 눈에 띄게 하는 것이 목적**이다 —
#: 검증되지 않은 LLM 가설이 기록된 사실과 섞여 보이면 안 된다.
ORIGIN_FILL = {
    TCOrigin.RECORDED: PatternFill("solid", fgColor="E8F5E9"),   # 연한 초록 = 관측됨
    TCOrigin.INFERRED: PatternFill("solid", fgColor="FFF4E5"),   # 연한 주황 = 추론(미검증)
    TCOrigin.USER: PatternFill("solid", fgColor="E3F2FD"),       # 연한 파랑 = 사람이 추가
}

PRIORITY_FONT = {
    "High": Font(bold=True, color="C62828"),
    "Medium": Font(color="333333"),
    "Low": Font(color="888888"),
}

_HEADER_FILL = PatternFill("solid", fgColor="37474F")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN = Side(style="thin", color="CFD8DC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center")


def export_excel(
    store: SessionStore,
    graph: FlowGraph,
    testcases: Sequence[TestCase],
    out_path: Path | str | None = None,
    *,
    embed_thumbnails: bool = True,
) -> Path:
    """TC를 xlsx로 내보낸다. 반환값은 생성된 파일 경로."""
    meta = store.get_session()
    path = Path(out_path) if out_path else store.export_dir / f"{meta.id}_TC.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _sheet_testcases(wb, store, testcases, embed_thumbnails)
    _sheet_coverage(wb, graph, testcases)
    _sheet_summary(wb, meta, graph, testcases)
    wb.save(path)
    return path


# ---------------------------------------------------------------- TC 시트


def _sheet_testcases(
    wb: Workbook, store: SessionStore, testcases: Sequence[TestCase], embed: bool
) -> None:
    ws = wb.active
    ws.title = "테스트케이스"
    _write_header(ws, HEADERS)

    thumb_dir = store.export_dir / "thumbs"
    if embed:
        thumb_dir.mkdir(parents=True, exist_ok=True)

    for row, tc in enumerate(testcases, start=2):
        # 절차와 기대결과는 번호를 붙여 셀 안에서 대응이 보이게 한다
        steps = "\n".join(f"{i}. {s}" for i, s in enumerate(tc.steps, 1))
        expected = "\n".join(f"{i}. {s}" for i, s in enumerate(tc.expected, 1))

        values = [
            tc.id,
            tc.category_major,
            tc.category_minor,
            tc.title,
            tc.precondition,
            steps,
            expected,
            tc.priority.value,
            tc.kind.value,
            tc.origin.value,
            "",  # 근거 = 썸네일/링크
            tc.rationale,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = _WRAP_TOP
            cell.border = _BORDER

        ws.cell(row=row, column=8).font = PRIORITY_FONT.get(tc.priority.value, Font())
        ws.cell(row=row, column=8).alignment = _CENTER
        ws.cell(row=row, column=9).alignment = _CENTER

        origin_cell = ws.cell(row=row, column=10)
        origin_cell.fill = ORIGIN_FILL.get(tc.origin, PatternFill())
        origin_cell.alignment = _CENTER
        if tc.origin is TCOrigin.INFERRED:
            origin_cell.font = Font(bold=True, color="E65100")

        height = _attach_evidence(ws, row, tc, store, thumb_dir, embed)
        ws.row_dimensions[row].height = height

    ws.freeze_panes = "A2"
    if testcases:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(testcases) + 1}"


def _attach_evidence(
    ws: Worksheet,
    row: int,
    tc: TestCase,
    store: SessionStore,
    thumb_dir: Path,
    embed: bool,
) -> float:
    """근거 칸에 썸네일을 넣고 원본 파일로 하이퍼링크를 건다.

    반환값은 이 행에 필요한 높이(포인트). 썸네일이 없으면 텍스트 기준 기본값.
    """
    cell = ws.cell(row=row, column=11)
    cell.border = _BORDER
    cell.alignment = _CENTER

    frame_id = next((f for f in tc.evidence_frames if store.frame(f)), None)
    if frame_id is None:
        cell.value = "(근거 없음)"
        cell.font = Font(color="999999", size=9)
        return 84.0

    frame = store.frame(frame_id)
    assert frame is not None
    src = store.frame_path(frame)

    # 하이퍼링크는 항상 건다 — 썸네일이 실패해도 원본은 열 수 있어야 한다
    cell.hyperlink = src.resolve().as_uri()
    cell.value = "원본 열기"
    cell.font = Font(color="1565C0", underline="single", size=9)

    if not embed:
        return 84.0

    thumb = thumb_dir / f"{frame_id}.png"
    if not thumb.exists():
        img = cv2.imread(str(src), cv2.IMREAD_COLOR)
        if img is None:
            return 84.0
        h, w = img.shape[:2]
        scale = THUMB_W / w
        small = cv2.resize(img, (THUMB_W, max(1, int(h * scale))), interpolation=cv2.INTER_AREA)
        cv2.imwrite(str(thumb), small)

    try:
        xl_img = XlImage(str(thumb))
        # 이미지가 셀을 덮으면 하이퍼링크를 못 누르므로 살짝 아래로 배치한다
        ws.add_image(xl_img, f"K{row}")
        return max(84.0, xl_img.height * 0.78)
    except Exception:
        return 84.0


# ---------------------------------------------------------------- 커버리지


def _sheet_coverage(wb: Workbook, graph: FlowGraph, testcases: Sequence[TestCase]) -> None:
    """어떤 TC도 커버하지 않은 전이 목록.

    **이 시트가 이 도구의 실질적 가치다.** TC를 자동 생성해주는 것보다,
    "이 경로는 아무도 테스트하지 않았다"를 명시적으로 알려주는 쪽이
    QA 실무에서 더 값지다.
    """
    ws = wb.create_sheet("커버리지")
    _write_header(
        ws,
        [
            ("전이 ID", 16),
            ("출발 화면", 24),
            ("행동", 30),
            ("도착 화면", 24),
            ("관측 횟수", 10),
            ("커버 여부", 12),
            ("커버한 TC", 26),
        ],
    )

    by_edge: dict[str, list[str]] = {}
    for tc in testcases:
        for eid in tc.edge_ids:
            by_edge.setdefault(eid, []).append(tc.id)
    covered, uncovered = coverage(graph, testcases)

    # 미커버를 위로 올려 눈에 먼저 들어오게 한다
    ordered = sorted(
        graph.ordered_transitions(), key=lambda t: (t.id in covered, t.from_state)
    )
    for row, t in enumerate(ordered, start=2):
        src = graph.states.get(t.from_state)
        dst = graph.states.get(t.to_state)
        is_covered = t.id in covered
        values = [
            t.id,
            src.name if src else t.from_state,
            t.action_desc,
            dst.name if dst else t.to_state,
            t.observed_count,
            "커버됨" if is_covered else "미커버",
            ", ".join(by_edge.get(t.id, [])),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = _WRAP_TOP
            cell.border = _BORDER
        status = ws.cell(row=row, column=6)
        if is_covered:
            status.fill = PatternFill("solid", fgColor="E8F5E9")
        else:
            status.fill = PatternFill("solid", fgColor="FFEBEE")
            status.font = Font(bold=True, color="C62828")
        status.alignment = _CENTER

    ws.freeze_panes = "A2"
    if ordered:
        ws.auto_filter.ref = f"A1:G{len(ordered) + 1}"
        rate = len(covered) / max(1, len(graph.transitions))
        note = ws.cell(
            row=len(ordered) + 3,
            column=1,
            value=f"전이 커버리지: {len(covered)}/{len(graph.transitions)} ({rate:.0%})"
            f"  ·  미커버 {len(uncovered)}건은 테스트되지 않은 경로입니다.",
        )
        note.font = Font(bold=True, size=11)


# ---------------------------------------------------------------- 요약


def _sheet_summary(
    wb: Workbook, meta: SessionMeta, graph: FlowGraph, testcases: Sequence[TestCase]
) -> None:
    ws = wb.create_sheet("요약")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 46

    row = 1

    def section(title: str) -> None:
        nonlocal row
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = _HEADER_FILL
        ws.cell(row=row, column=2).fill = _HEADER_FILL
        row += 1

    def item(key: str, value: object) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(value)).alignment = _WRAP_TOP
        row += 1

    section("세션 정보")
    item("세션 ID", meta.id)
    item("게임", meta.game_name)
    item("녹화 시작", meta.started_at)
    item("녹화 종료", meta.ended_at or "(미종료)")
    item("캡처 백엔드", meta.capture_backend or "(미기록)")
    item("해상도", f"{meta.client_w}x{meta.client_h}")
    if meta.note:
        item("메모", meta.note)
    row += 1

    section("분석 결과")
    item("화면 상태", f"{len(graph.states)}개 (표시 {len(graph.visible_states())}개)")
    item("전이", f"{len(graph.transitions)}개")
    covered, uncovered = coverage(graph, testcases)
    item(
        "전이 커버리지",
        f"{len(covered)}/{len(graph.transitions)}"
        f" ({len(covered) / max(1, len(graph.transitions)):.0%}) · 미커버 {len(uncovered)}건",
    )
    row += 1

    section("테스트케이스")
    item("총 건수", f"{len(testcases)}건")
    for label, counter in (
        ("출처별", Counter(tc.origin.value for tc in testcases)),
        ("유형별", Counter(tc.kind.value for tc in testcases)),
        ("우선순위별", Counter(tc.priority.value for tc in testcases)),
        ("대분류별", Counter(tc.category_major for tc in testcases)),
    ):
        item(label, ", ".join(f"{k} {v}건" for k, v in counter.most_common()) or "(없음)")
    row += 1

    section("읽는 방법")
    item(
        "출처 = 기록됨",
        "실제 플레이에서 관측된 경로입니다. 근거 스크린샷이 있습니다.",
    )
    item(
        "출처 = 추론됨",
        "화면에 보이는 UI 요소에서 LLM이 추론한 미관측 케이스입니다. "
        "실제로 그렇게 동작하는지는 검증되지 않았으므로, 수행 전 타당성을 확인하세요.",
    )
    item(
        "커버리지 시트",
        "어떤 TC도 커버하지 않은 전이 목록입니다. 테스트 공백을 여기서 확인하세요.",
    )


# ---------------------------------------------------------------- 공통


def _write_header(ws: Worksheet, headers: Iterable[tuple[str, int]]) -> None:
    for col, (title, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 24
