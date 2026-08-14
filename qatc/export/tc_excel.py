"""컨텐츠 지식에서 만든 TC를 xlsx로 내보낸다.

녹화 파이프라인의 출력기(그래프·세션 저장소가 필요했다)와 달리, 이 모듈은
지식 저장소(:class:`~qatc.knowledge.store.KnowledgeStore`)만 있으면 된다.
녹화 세션이 없는 인터뷰 기반 파이프라인의 출력 경로다.

**모든 셀 값은 :func:`clean_cell` 을 통과시킨다.** openpyxl은 제어문자가 든
문자열을 거부하는데, 예전 구현에 이 방어가 없어 OCR·키 이름에 섞여 들어온
`\\x03` 하나가 export 전체를 죽였다.
"""

from __future__ import annotations

import re
from collections import Counter
from pathlib import Path
from typing import Collection, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..knowledge.gate import FamilySkip
from ..models import TCOrigin, TestCase

#: 근거가 살아 있는 TC / 철회된 TC 의 표시. 빈 칸은 "누락"으로 읽히므로 둘 다 쓴다.
_EVIDENCE_LIVE = "유효"
_EVIDENCE_WITHDRAWN = "근거 철회됨"

#: openpyxl이 거부하는 제어문자. 탭·개행·복귀는 남긴다 (셀 안에서 유효하다).
_ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_WRAP = Alignment(vertical="top", wrap_text=True)

_ORIGIN_FILL = {
    TCOrigin.INTERVIEW: PatternFill("solid", fgColor="E2EFDA"),
    TCOrigin.INFERRED: PatternFill("solid", fgColor="FCE4D6"),
    TCOrigin.USER: PatternFill("solid", fgColor="DDEBF7"),
    TCOrigin.RECORDED: PatternFill("solid", fgColor="E2EFDA"),
}


def clean_cell(value: str) -> str:
    """셀에 넣어도 안전한 문자열로 만든다."""
    return _ILLEGAL.sub("", value)


def _header(ws, titles: Sequence[str], widths: Sequence[int]) -> None:
    for i, (t, w) in enumerate(zip(titles, widths), start=1):
        c = ws.cell(row=1, column=i, value=t)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _WRAP
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _sheet_testcases(
    wb: Workbook, cases: Sequence[TestCase], withdrawn: Collection[str]
) -> None:
    ws = wb.active
    ws.title = "테스트케이스"
    _header(
        ws,
        ["TC ID", "대분류", "중분류", "제목", "사전조건", "절차", "기대결과",
         "우선순위", "유형", "출처", "근거", "근거 상태"],
        [14, 14, 14, 40, 28, 40, 40, 10, 10, 10, 36, 14],
    )
    for r, tc in enumerate(cases, start=2):
        values = [
            tc.id, tc.category_major, tc.category_minor, tc.title, tc.precondition,
            "\n".join(f"{i}. {s}" for i, s in enumerate(tc.steps, 1)),
            "\n".join(f"- {s}" for s in tc.expected),
            tc.priority.value, tc.kind.value, tc.origin.value, tc.rationale,
            _EVIDENCE_WITHDRAWN if tc.category_minor in withdrawn else _EVIDENCE_LIVE,
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=clean_cell(str(v)))
            cell.alignment = _WRAP
        ws.cell(row=r, column=10).fill = _ORIGIN_FILL.get(tc.origin, PatternFill())


def _sheet_skipped(
    wb: Workbook, skipped: Sequence[FamilySkip], withdrawn: Collection[str]
) -> None:
    """미확인 슬롯 시트.

    예전 헤더는 `만들지 못한 계열` 이었는데, 근거가 나중에 철회된 계열에 대해서는
    **거짓이다** — 같은 통합문서의 `테스트케이스` 시트에 그 계열 TC 가 버젓이 있다.
    계열 이름만 적고, TC 가 실제로 있는지는 별도 칸으로 정직하게 밝힌다.
    """
    ws = wb.create_sheet("미확인 항목")
    _header(ws, ["슬롯", "묻는 것", "계열", "상태", "이 계열의 TC"],
            [18, 40, 18, 40, 34])
    for r, s in enumerate(skipped, start=2):
        tc_state = (f"{_EVIDENCE_WITHDRAWN} — 이전에 만든 TC가 남아 있습니다"
                    if s.family in withdrawn else "TC 없음")
        for col, v in enumerate(
            [s.slot_key, s.prompt_hint, s.family, s.reason, tc_state], start=1
        ):
            ws.cell(row=r, column=col, value=clean_cell(str(v))).alignment = _WRAP


def _sheet_summary(
    wb: Workbook,
    content: str,
    cases: Sequence[TestCase],
    skipped: Sequence[FamilySkip],
    withdrawn: Collection[str],
) -> None:
    ws = wb.create_sheet("요약")
    _header(ws, ["항목", "값"], [28, 60])

    by_origin = Counter(tc.origin.value for tc in cases)
    by_kind = Counter(tc.kind.value for tc in cases)
    rows = [
        ("컨텐츠", content),
        ("TC 총계", str(len(cases))),
        ("유형별", ", ".join(f"{k} {v}" for k, v in sorted(by_kind.items())) or "-"),
        ("출처별", ", ".join(f"{k} {v}" for k, v in sorted(by_origin.items())) or "-"),
        ("미확인 항목", str(len(skipped))),
        ("근거 철회된 TC",
         str(sum(1 for tc in cases if tc.category_minor in withdrawn))),
        ("읽는 방법",
         "출처 '인터뷰'는 담당자가 진술한 내용, '추론됨'은 진술에서 도출한 것입니다. "
         "'추론됨'은 실제로 그렇게 동작하는지 검증되지 않았습니다. "
         "'미확인 항목' 시트는 아직 설명되지 않은 항목과 그 계열입니다. "
         "'근거 상태'가 '근거 철회됨'인 TC는 만든 뒤에 근거 슬롯이 다시 열린 "
         "것으로, 지우지 않고 남겨 두었습니다 — 내용을 검토한 뒤 채택 여부를 정하세요."),
    ]
    for r, (k, v) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=clean_cell(k)).alignment = _WRAP
        ws.cell(row=r, column=2, value=clean_cell(v)).alignment = _WRAP


class ExportBlocked(OSError):
    """대상 파일에 쓸 수 없다 — 보통 Excel 이 그 파일을 열어 두고 있다.

    메시지는 완성된 한국어 문장이라 호출자는 ``str(exc)`` 를 그대로 쓰면 된다.
    CLI 와 앱이 같은 문장을 쓰게 하려고 CLI 가 아니라 여기서 만든다.
    """


def export_tc_excel(
    content: str,
    testcases: Sequence[TestCase],
    skipped: Sequence[FamilySkip],
    out_path: Path | str,
    withdrawn: Collection[str] = (),
) -> Path:
    """TC를 xlsx로 내보낸다. 반환값은 생성된 파일 경로.

    :param withdrawn: 근거가 철회된 계열 이름
        (:func:`~qatc.knowledge.gate.withdrawn_families`). TC를 **지우지 않고**
        표시만 하므로, 이 값이 비어 있으면 산출물이 자기모순 상태로 나간다.
    """
    path = Path(out_path)

    wb = Workbook()
    _sheet_testcases(wb, testcases, withdrawn)
    _sheet_skipped(wb, skipped, withdrawn)
    _sheet_summary(wb, content, testcases, skipped, withdrawn)
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
    except PermissionError as exc:
        raise ExportBlocked(
            f"파일에 쓸 수 없습니다 — {path}. "
            f"Excel에서 이 파일을 닫고 다시 시도하세요."
        ) from exc
    return path
